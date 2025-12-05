#!/usr/bin/env python3
"""
Convert F5 As-Built JSON into an Excel workbook.

Input JSON is expected to be the output from f5_asbuilt.py --format json, with keys:

- device_report
- virtual_servers
- pools
- nodes
- monitors
- irules
- ssl_profiles
- certificates
- usage (irule_usage, monitor_usage, ssl_profile_usage, cert_usage)

The Excel workbook will contain sheets:

1. Virtual_Servers
2. Pools
3. Nodes
4. Monitors
5. IRules
6. SSL_Profiles

By default, the Excel filename will be the JSON filename with extension changed to .xlsx.
"""

import argparse
import json
import os
import sys
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# ----------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------


def load_json(path: str) -> Dict[str, Any]:
    if not os.path.exists(path):
        print(f"[ERROR] JSON file not found: {path}", file=sys.stderr)
        sys.exit(1)
    with open(path, "r", encoding="utf-8") as f:
        try:
            return json.load(f)
        except json.JSONDecodeError as e:
            print(f"[ERROR] Failed to parse JSON: {e}", file=sys.stderr)
            sys.exit(1)


def default_excel_name(json_path: str) -> str:
    base, _ = os.path.splitext(json_path)
    # Modern Excel format; rename to .xls if you really need legacy extension
    return base + ".xlsx"


# ----------------------------------------------------------------------
# Sheet builders
# ----------------------------------------------------------------------


def build_virtual_servers_sheet(
    ws: Worksheet, virtual_servers: List[Dict[str, Any]]
) -> None:
    ws.title = "Virtual_Servers"
    headers = [
        "Name",
        "IP",
        "Port",
        "Pool",
        "Profiles",
        "Persistence",
        "iRules",
    ]
    ws.append(headers)

    for vs in virtual_servers:
        profiles = ", ".join(vs.get("profiles") or [])
        persistence = ", ".join(vs.get("persistence") or [])
        irules = ", ".join(vs.get("irules") or [])
        ws.append(
            [
                vs.get("name"),
                vs.get("destination_ip"),
                vs.get("destination_port"),
                vs.get("pool"),
                profiles,
                persistence,
                irules,
            ]
        )


def build_pools_sheet(ws: Worksheet, pools: List[Dict[str, Any]]) -> None:
    ws.title = "Pools"
    headers = [
        "Pool_Name",
        "LB_Method",
        "Monitor",
        "Member_Name",
        "Member_Address",
        "Member_State",
        "Member_Session",
    ]
    ws.append(headers)

    for p in pools:
        pool_name = p.get("name")
        lb_method = p.get("lb_method")
        monitor = p.get("monitor")
        members = p.get("members") or []

        if not members:
            # Pool with no members still gets one row
            ws.append([pool_name, lb_method, monitor, None, None, None, None])
            continue

        for m in members:
            ws.append(
                [
                    pool_name,
                    lb_method,
                    monitor,
                    m.get("name"),
                    m.get("address"),
                    m.get("state"),
                    m.get("session"),
                ]
            )


def build_nodes_sheet(ws: Worksheet, nodes: List[Dict[str, Any]]) -> None:
    ws.title = "Nodes"
    headers = ["Node_Name", "IP_Address", "State", "Session"]
    ws.append(headers)

    for n in nodes:
        ws.append(
            [
                n.get("name"),
                n.get("address"),
                n.get("state"),
                n.get("session"),
            ]
        )


def build_monitors_sheet(
    ws: Worksheet,
    monitors: List[Dict[str, Any]],
    usage: Dict[str, Any],
) -> None:
    ws.title = "Monitors"
    monitor_usage: Dict[str, List[str]] = usage.get("monitor_usage", {}) or {}
    headers = ["Monitor_Name", "Type", "Partition", "Used_By_Pools"]
    ws.append(headers)

    for m in sorted(monitors, key=lambda x: x.get("name") or ""):
        name = m.get("name")
        used_by = ", ".join(monitor_usage.get(name, []))
        ws.append(
            [
                name,
                m.get("type"),
                m.get("partition"),
                used_by,
            ]
        )


def build_irules_sheet(
    ws: Worksheet,
    irules: List[Dict[str, Any]],
    usage: Dict[str, Any],
) -> None:
    ws.title = "IRules"
    irule_usage: Dict[str, List[str]] = usage.get("irule_usage", {}) or {}
    headers = ["IRule_Name", "Partition", "Used_By_Virtual_Servers"]
    ws.append(headers)

    for r in sorted(irules, key=lambda x: x.get("name") or ""):
        name = r.get("name")
        used_by = ", ".join(irule_usage.get(name, []))
        ws.append(
            [
                name,
                r.get("partition"),
                used_by,
            ]
        )


def build_ssl_profiles_sheet(
    ws: Worksheet,
    ssl_profiles: List[Dict[str, Any]],
    certificates: List[Dict[str, Any]],
    usage: Dict[str, Any],
) -> None:
    ws.title = "SSL_Profiles"
    ssl_usage: Dict[str, List[str]] = usage.get("ssl_profile_usage", {}) or {}

    # Build map: cert_name -> expiration
    cert_exp_map: Dict[str, str] = {}
    for c in certificates:
        # c["name"] is typically the object name; we match on last path part of profile["cert"]
        cert_name = c.get("name")
        cert_exp = c.get("expiration")
        if cert_name:
            cert_exp_map[cert_name] = cert_exp

    headers = [
        "Profile_Name",
        "Partition",
        "Certificate",
        "Certificate_Expiration",
        "Attached_Virtual_Servers",
    ]
    ws.append(headers)

    for sp in sorted(ssl_profiles, key=lambda x: x.get("name") or ""):
        name = sp.get("name")
        cert_path = sp.get("cert")
        cert_name = None
        if cert_path:
            cert_name = cert_path.split("/")[-1]

        expiration = cert_exp_map.get(cert_name, None) if cert_name else None
        attached_vips = ", ".join(ssl_usage.get(name, []))

        ws.append(
            [
                name,
                sp.get("partition"),
                cert_path,
                expiration,
                attached_vips,
            ]
        )


# ----------------------------------------------------------------------
# Main orchestration
# ----------------------------------------------------------------------


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="F5 As-Built JSON → Excel generator")
    parser.add_argument(
        "json_file",
        help="Path to JSON file generated by f5_asbuilt.py --format json",
    )
    parser.add_argument(
        "-o",
        "--output",
        help="Output Excel filename (default: same as JSON with .xlsx)",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    # Resolve JSON path:
    # 1) if args.json_file exists as-is, use it
    # 2) otherwise, try ./json/<args.json_file>
    json_path = args.json_file
    if not os.path.exists(json_path):
        candidate = os.path.join("json", json_path)
        if os.path.exists(candidate):
            json_path = candidate
        else:
            print(
                f"[ERROR] JSON file not found: {args.json_file} or {candidate}",
                file=sys.stderr,
            )
            sys.exit(1)

    data = load_json(json_path)

    device_report = data.get("device_report", {})
    virtual_servers = data.get("virtual_servers", [])
    pools = data.get("pools", [])
    nodes = data.get("nodes", [])
    monitors = data.get("monitors", [])
    irules = data.get("irules", [])
    ssl_profiles = data.get("ssl_profiles", [])
    certificates = data.get("certificates", [])
    usage = data.get("usage", {}) or {}

    # Create workbook and sheets
    wb = Workbook()
    # Default sheet becomes Virtual_Servers
    ws_vs = wb.active
    build_virtual_servers_sheet(ws_vs, virtual_servers)

    ws_pools = wb.create_sheet(title="Pools")
    build_pools_sheet(ws_pools, pools)

    ws_nodes = wb.create_sheet(title="Nodes")
    build_nodes_sheet(ws_nodes, nodes)

    ws_mon = wb.create_sheet(title="Monitors")
    build_monitors_sheet(ws_mon, monitors, usage)

    ws_irules = wb.create_sheet(title="IRules")
    build_irules_sheet(ws_irules, irules, usage)

    ws_ssl = wb.create_sheet(title="SSL_Profiles")
    build_ssl_profiles_sheet(ws_ssl, ssl_profiles, certificates, usage)

    # Determine output filename + default XLS directory
    if args.output:
        out_path = args.output
    else:
        # base name only (strip json/ and strip .json)
        base = os.path.basename(json_path)
        out_path = default_excel_name(base)

        # put into ./xls folder
        out_dir = "xls"
        os.makedirs(out_dir, exist_ok=True)
        out_path = os.path.join(out_dir, out_path)

    wb.save(out_path)
    print(f"Wrote Excel workbook to: {out_path}")


if __name__ == "__main__":
    main()
